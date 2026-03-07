import os
import io
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta, time
import pytz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import config

app = Flask(__name__)
app.config['SECRET_KEY'] = config.SECRET_KEY
app.config['UPLOAD_FOLDER'] = 'uploads'

# --- CONFIGURACIÓN DE BASE DE DATOS (MYSQL) ---
app.config['SQLALCHEMY_DATABASE_URI'] = config.SQLALCHEMY_DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True, 'pool_recycle': 280}

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# --- MODELOS DE DATOS ---
class User(UserMixin, db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    biometric_id = db.Column(db.String(20), unique=True)
    nombre = db.Column(db.String(100))
    username = db.Column(db.String(50), unique=True)
    password = db.Column(db.String(255))
    rol = db.Column(db.String(20), default='docente')
    acceso_puerta = db.Column(db.Integer, default=0)

class Log(db.Model):
    __tablename__ = 'logs'
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, nullable=False)
    usuario_id = db.Column(db.String(20))
    tipo_evento = db.Column(db.String(50))
    origen = db.Column(db.String(50))
    latitud = db.Column(db.Float, nullable=True)
    longitud = db.Column(db.Float, nullable=True)
    descripcion = db.Column(db.Text, nullable=True)
    foto_path = db.Column(db.String(255), nullable=True)

class Comando(db.Model):
    __tablename__ = 'comandos'
    id = db.Column(db.Integer, primary_key=True)
    instruccion = db.Column(db.String(100))
    estado = db.Column(db.String(20), default='PENDIENTE')

class Permiso(db.Model):
    __tablename__ = 'permisos'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    fecha_permiso = db.Column(db.Date, nullable=False)
    observacion = db.Column(db.Text, nullable=True)
    docente = db.relationship('User', backref=db.backref('permisos', lazy=True))

# --- INICIALIZACIÓN ---
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def init_db():
    with app.app_context():
        db.create_all()
        # Crear carpeta de uploads si no existe
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
        
        if not User.query.filter_by(username='admin').first():
            hashed_pw = generate_password_hash('istae123A*', method='pbkdf2:sha256')
            admin = User(biometric_id='999', nombre='Admin Principal', username='admin', password=hashed_pw, rol='admin', acceso_puerta=1)
            db.session.add(admin)
            db.session.commit()

init_db()

# --- API PARA HARDWARE (IOT) ---
@app.route('/api/sincronizar')
def api_sincronizar():
    usuarios = User.query.filter_by(acceso_puerta=1).all()
    return ",".join([str(u.biometric_id) for u in usuarios])

@app.route('/api/recibir_log', methods=['POST'])
def api_recibir_log():
    data = request.json
    if not data or data.get('token') != config.TOKEN_NODE:
        return jsonify({"status": "error", "message": "Token inválido"}), 403

    fecha_log = None
    fecha_str = data.get('fecha_dispositivo')
    tz_ecu = pytz.timezone('America/Guayaquil')

    if fecha_str:
        try:
            fecha_limpia = fecha_str.replace('T', ' ')
            fecha_log = datetime.strptime(fecha_limpia[:19], "%Y-%m-%d %H:%M:%S")
        except:
            fecha_log = None

    if fecha_log is None:
        fecha_log = datetime.now(tz_ecu)

    nuevo_log = Log(
        fecha=fecha_log,
        usuario_id=data.get('id'),
        tipo_evento="Asistencia + puerta",
        origen="Huella"
    )
    db.session.add(nuevo_log)
    db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/check_comando')
def api_check_comando():
    cmd = Comando.query.filter_by(estado='PENDIENTE').order_by(Comando.id.asc()).first()
    if cmd:
        cmd.estado = 'ENVIADO'
        db.session.commit()
        return cmd.instruccion
    return "NADA"

# --- VISTAS Y DASHBOARDS ---
@app.route('/')
@login_required
def index():
    return redirect(url_for('admin_dashboard' if current_user.rol == 'admin' else 'docente_dashboard'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.rol != 'admin': return redirect(url_for('docente_dashboard'))
    docentes = User.query.filter_by(rol='docente').all()
    return render_template('admin.html', docentes=docentes)

@app.route('/docente/dashboard')
@login_required
def docente_dashboard():
    logs = Log.query.filter_by(usuario_id=current_user.biometric_id).order_by(Log.id.desc()).limit(10).all()
    return render_template('docente.html', logs=logs)

@app.route('/perfil')
@login_required
def perfil():
    return render_template('perfil.html')

# --- RUTA PARA SERVIR EVIDENCIAS ---
@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# --- ACCIONES ---
@app.route('/admin/abrir')
@login_required
def admin_abrir():
    db.session.add(Comando(instruccion='ABRIR'))
    db.session.add(Log(
        fecha=datetime.now(pytz.timezone('America/Guayaquil')),
        usuario_id=current_user.biometric_id,
        tipo_evento="Apertura Remota",
        origen="Panel Control"
    ))
    db.session.commit()
    flash('Comando de apertura enviado.', 'danger')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/sincronizar_hora', methods=['POST'])
@login_required
def admin_sincronizar_hora():
    if current_user.rol != 'admin': return redirect(url_for('index'))
    
    time_str = request.form.get('new_time')
    if time_str:
        try:
            dt_obj = datetime.strptime(time_str, '%Y-%m-%dT%H:%M')
            iso_time = dt_obj.strftime('%Y-%m-%dT%H:%M:%S-05:00')
            
            comando_str = f"SET_TIME|{iso_time}"
            db.session.add(Comando(instruccion=comando_str))
            db.session.commit()
            flash(f'Comando de sincronización de hora ({iso_time}) enviado al dispositivo.', 'info')
        except ValueError:
            flash('Formato de fecha y hora inválido.', 'danger')

    return redirect(url_for('admin_dashboard') + '#collapseSync')

@app.route('/docente/abrir_puerta')
@login_required
def docente_abrir():
    if current_user.acceso_puerta == 1:
        db.session.add(Comando(instruccion='ABRIR'))
        db.session.add(Log(
            fecha=datetime.now(pytz.timezone('America/Guayaquil')),
            usuario_id=current_user.biometric_id,
            tipo_evento="Apertura Remota",
            origen="Asistencia remota"
        ))
        db.session.commit()
        flash('Puerta abierta.', 'success')
    return redirect(url_for('docente_dashboard'))

@app.route('/docente/marcar_web', methods=['GET', 'POST'])
@login_required
def docente_marcar():
    if request.method == 'POST':
        lat = request.form.get('latitud')
        lon = request.form.get('longitud')
        descripcion = request.form.get('descripcion')
        foto = request.files.get('foto')

        # Validación básica de GPS
        if not lat or not lon:
            flash('No se pudo obtener la ubicación GPS. Intente de nuevo.', 'danger')
            return redirect(url_for('docente_dashboard'))

        filename = None
        if foto and foto.filename != '':
            filename = secure_filename(f"{current_user.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{foto.filename}")
            foto.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        nuevo_log = Log(
            fecha=datetime.now(pytz.timezone('America/Guayaquil')),
            usuario_id=current_user.biometric_id,
            tipo_evento="Asistencia",
            origen="Asistencia remota",
            latitud=float(lat),
            longitud=float(lon),
            descripcion=descripcion,
            foto_path=filename
        )
        db.session.add(nuevo_log)
        db.session.commit()
        flash('Asistencia remota registrada con éxito.', 'success')
        return redirect(url_for('docente_dashboard'))

    # Si es GET, simplemente renderiza la misma página (la lógica del modal se encarga)
    return redirect(url_for('docente_dashboard'))

# --- GESTIÓN DE ASISTENCIAS Y PERMISOS ---

@app.route('/admin/gestion_asistencia', methods=['GET'])
@login_required
def gestion_asistencia():
    if current_user.rol != 'admin':
        return redirect(url_for('index'))

    fecha_ini_str = request.args.get('fecha_inicio')
    fecha_fin_str = request.args.get('fecha_fin')
    docente_id_filtro = request.args.get('docente_id')

    query = db.session.query(Log, User).outerjoin(User, Log.usuario_id == User.biometric_id).order_by(Log.fecha.desc())

    if fecha_ini_str:
        start_dt = datetime.strptime(fecha_ini_str, '%Y-%m-%d')
        query = query.filter(Log.fecha >= start_dt)
    if fecha_fin_str:
        end_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
        query = query.filter(Log.fecha <= end_dt + timedelta(days=1))
    if docente_id_filtro and docente_id_filtro != 'todos':
        # Assuming docente_id in filter is the User ID
        user_filter = db.session.get(User, docente_id_filtro)
        if user_filter:
            query = query.filter(Log.usuario_id == user_filter.biometric_id)

    logs_data = query.all()
    docentes = User.query.filter_by(rol='docente').all()

    return render_template('gestion_asistencia.html', 
                           logs_data=logs_data, 
                           docentes=docentes,
                           filtros=request.args
                           )

@app.route('/admin/asistencia/editar/<int:id>', methods=['GET'])
@login_required
def editar_asistencia(id):
    if current_user.rol != 'admin': return redirect(url_for('index'))
    
    log = db.session.get(Log, id)
    if not log:
        flash('El registro de asistencia no existe.', 'danger')
        return redirect(url_for('gestion_asistencia'))

    docentes = User.query.filter_by(rol='docente').all()
    log_user = User.query.filter_by(biometric_id=log.usuario_id).first()

    return render_template('editar_asistencia.html', log=log, docentes=docentes, log_user=log_user)

@app.route('/admin/asistencia/actualizar', methods=['POST'])
@login_required
def actualizar_asistencia():
    if current_user.rol != 'admin': return redirect(url_for('index'))

    log_id = request.form.get('log_id')
    log = db.session.get(Log, log_id)

    if not log:
        flash('El registro no existe.', 'danger')
        return redirect(url_for('gestion_asistencia'))

    log.fecha = datetime.strptime(request.form['fecha'], '%Y-%m-%dT%H:%M')
    user = db.session.get(User, request.form['docente_id'])
    if user:
        log.usuario_id = user.biometric_id
        
    log.tipo_evento = request.form['tipo_evento']
    log.origen = request.form['origen']
    log.descripcion = request.form['descripcion']
    
    db.session.commit()
    flash('Registro de asistencia actualizado correctamente.', 'success')
    return redirect(url_for('gestion_asistencia'))

@app.route('/admin/asistencia/eliminar/<int:id>')
@login_required
def eliminar_asistencia(id):
    if current_user.rol != 'admin': return redirect(url_for('index'))

    log = db.session.get(Log, id)
    if log:
        if log.foto_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], log.foto_path))
            except OSError as e:
                flash(f'Error al eliminar la foto de evidencia: {e}', 'warning')
        
        db.session.delete(log)
        db.session.commit()
        flash('Registro de asistencia eliminado.', 'success')
    else:
        flash('El registro no existe.', 'danger')
    
    return redirect(url_for('gestion_asistencia'))

@app.route('/admin/gestion_permisos', methods=['GET'])
@login_required
def gestion_permisos():
    if current_user.rol != 'admin':
        return redirect(url_for('index'))

    fecha_ini_str = request.args.get('fecha_inicio')
    fecha_fin_str = request.args.get('fecha_fin')
    docente_id_filtro = request.args.get('docente_id')

    query = Permiso.query.join(User).order_by(Permiso.fecha_permiso.desc())

    if fecha_ini_str:
        start_dt = datetime.strptime(fecha_ini_str, '%Y-%m-%d').date()
        query = query.filter(Permiso.fecha_permiso >= start_dt)
    if fecha_fin_str:
        end_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        query = query.filter(Permiso.fecha_permiso <= end_dt)
    if docente_id_filtro and docente_id_filtro != 'todos':
        query = query.filter(Permiso.user_id == docente_id_filtro)

    permisos = query.all()
    docentes = User.query.filter_by(rol='docente').all()

    return render_template('gestion_permisos.html', 
                           permisos=permisos, 
                           docentes=docentes,
                           filtros=request.args
                           )

# --- GESTIÓN DOCENTES ---
@app.route('/crear_docente', methods=['POST'])
@login_required
def crear_docente():
    hashed = generate_password_hash(request.form['password'], method='pbkdf2:sha256')
    db.session.add(User(
        biometric_id=request.form['bio_id'],
        nombre=request.form['nombre'],
        username=request.form['username'],
        password=hashed,
        acceso_puerta=1 if request.form.get('acceso_puerta') else 0
    ))
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/actualizar_docente', methods=['POST'])
@login_required
def actualizar_docente():
    u = db.session.get(User, request.form.get('user_id'))
    u.nombre = request.form['nombre']
    u.biometric_id = request.form['bio_id']
    u.username = request.form['username']
    u.acceso_puerta = 1 if request.form.get('acceso_puerta') else 0
    if request.form.get('password'):
        u.password = generate_password_hash(request.form['password'], method='pbkdf2:sha256')
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/eliminar_docente/<int:id>')
@login_required
def eliminar_docente(id):
    u = db.session.get(User, id)
    if u and u.username != 'admin':
        db.session.delete(u)
        db.session.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/editar_docente/<int:id>')
@login_required
def editar_docente(id):
    u = db.session.get(User, id)
    return render_template('editar_docente.html', docente=u)
    
# --- GESTIÓN DE PERMISOS ---
@app.route('/admin/permiso/crear', methods=['POST'])
@login_required
def crear_permiso():
    if current_user.rol != 'admin':
        return redirect(url_for('index'))

    docente_id = request.form.get('docente_id')
    fecha_str = request.form.get('fecha_permiso')
    observacion = request.form.get('observacion')

    if not docente_id or not fecha_str:
        flash('El docente y la fecha son obligatorios.', 'danger')
        return redirect(url_for('admin_dashboard') + '#collapsePermiso')

    fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()

    nuevo_permiso = Permiso(
        user_id=docente_id,
        fecha_permiso=fecha_obj,
        observacion=observacion
    )
    db.session.add(nuevo_permiso)
    db.session.commit()

    flash('Permiso registrado correctamente.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/permiso/editar/<int:id>', methods=['GET'])
@login_required
def editar_permiso(id):
    if current_user.rol != 'admin':
        return redirect(url_for('index'))
    
    permiso = db.session.get(Permiso, id)
    if not permiso:
        flash('El permiso no existe.', 'danger')
        return redirect(url_for('gestion_permisos'))

    docentes = User.query.filter_by(rol='docente').all()
    return render_template('editar_permiso.html', permiso=permiso, docentes=docentes)


@app.route('/admin/permiso/actualizar', methods=['POST'])
@login_required
def actualizar_permiso():
    if current_user.rol != 'admin':
        return redirect(url_for('index'))

    permiso_id = request.form.get('permiso_id')
    permiso = db.session.get(Permiso, permiso_id)

    if not permiso:
        flash('El permiso no existe.', 'danger')
        return redirect(url_for('gestion_permisos'))

    docente_id = request.form.get('docente_id')
    fecha_str = request.form.get('fecha_permiso')
    observacion = request.form.get('observacion')

    if not docente_id or not fecha_str:
        flash('El docente y la fecha son obligatorios.', 'danger')
        return redirect(url_for('editar_permiso', id=permiso_id))

    permiso.user_id = docente_id
    permiso.fecha_permiso = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    permiso.observacion = observacion
    
    db.session.commit()
    flash('Permiso actualizado correctamente.', 'success')
    return redirect(url_for('gestion_permisos'))


@app.route('/admin/permiso/eliminar/<int:id>')
@login_required
def eliminar_permiso(id):
    if current_user.rol != 'admin':
        return redirect(url_for('index'))

    permiso = db.session.get(Permiso, id)
    if permiso:
        db.session.delete(permiso)
        db.session.commit()
        flash('Permiso eliminado correctamente.', 'success')
    else:
        flash('El permiso no existe.', 'danger')
    
    return redirect(url_for('gestion_permisos'))

# --- REPORTES ---
@app.route('/descargar_reporte_permisos')
@login_required
def descargar_reporte_permisos():
    if current_user.rol != 'admin':
        return redirect(url_for('index'))

    fecha_ini_str = request.args.get('fecha_inicio_permiso')
    fecha_fin_str = request.args.get('fecha_fin_permiso')
    docente_id_filtro = request.args.get('docente_id_permiso')

    query = Permiso.query.join(User).order_by(Permiso.fecha_permiso.desc())

    if fecha_ini_str:
        start_dt = datetime.strptime(fecha_ini_str, '%Y-%m-%d').date()
        query = query.filter(Permiso.fecha_permiso >= start_dt)
    if fecha_fin_str:
        end_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        query = query.filter(Permiso.fecha_permiso <= end_dt)
    if docente_id_filtro and docente_id_filtro != 'todos':
        query = query.filter(Permiso.user_id == docente_id_filtro)

    permisos = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Permisos"

    fill_header = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_cell = ws.cell(row=1, column=1, value="Reporte de Permisos de Docentes")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = align_center

    headers = ["Docente", "Fecha del Permiso", "Observación"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.border = border_thin
        cell.alignment = align_center

    row_idx = 3
    for permiso in permisos:
        ws.cell(row=row_idx, column=1, value=permiso.docente.nombre).border = border_thin
        ws.cell(row=row_idx, column=2, value=permiso.fecha_permiso.strftime('%d/%m/%Y')).border = border_thin
        ws.cell(row=row_idx, column=3, value=permiso.observacion or 'N/A').border = border_thin
        ws.cell(row=row_idx, column=2).alignment = align_center
        row_idx += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        download_name=f"Reporte_Permisos_ISTAE_{datetime.now().strftime('%Y%m%d')}.xlsx",
        as_attachment=True
    )

@app.route('/descargar_reporte_matricial')
@login_required
def descargar_reporte_matricial():
    if current_user.rol != 'admin': return redirect(url_for('index'))
    
    # --- Filtros de Fecha y Docente ---
    fecha_ini_str = request.args.get('fecha_inicio')
    fecha_fin_str = request.args.get('fecha_fin')
    docente_filtro = request.args.get('docente_id')
    
    if fecha_ini_str and fecha_fin_str:
        start_dt = datetime.strptime(fecha_ini_str, '%Y-%m-%d')
        end_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
    else:
        end_dt = datetime.now()
        start_dt = end_dt - timedelta(days=6)

    dias_reporte = [ (start_dt + timedelta(days=x)).strftime('%Y-%m-%d') for x in range((end_dt-start_dt).days + 1) ]

    # --- Filtros de Hora ---
    try:
        hora_inicio_m = datetime.strptime(request.args.get('hora_inicio_m', '07:00'), '%H:%M').time()
        hora_fin_m = datetime.strptime(request.args.get('hora_fin_m', '13:00'), '%H:%M').time()
        hora_inicio_t = datetime.strptime(request.args.get('hora_inicio_t', '13:01'), '%H:%M').time()
        hora_fin_t = datetime.strptime(request.args.get('hora_fin_t', '22:00'), '%H:%M').time()
    except ValueError:
        flash('Formato de hora inválido. Use HH:MM.', 'danger')
        return redirect(url_for('admin_dashboard') + '#collapseReportes')

    # --- Consulta de Datos ---
    docentes = User.query.filter_by(biometric_id=docente_filtro).all() if docente_filtro and docente_filtro != 'todos' else User.query.filter_by(rol='docente').all()

    all_logs = Log.query.filter(
        Log.fecha >= start_dt,
        Log.fecha <= end_dt + timedelta(days=1),
        Log.tipo_evento.like('%Asistencia%')
    ).all()

    # --- Generación de Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Matricial"

    fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(dias_reporte))
    ws.cell(row=1, column=1, value=f"Reporte de Asistencia ({start_dt.strftime('%d/%m/%Y')} al {end_dt.strftime('%d/%m/%Y')})").font = Font(bold=True)

    headers = ["ID", "Nombre", "Depto"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill, c.border, c.alignment, c.font = fill_green, border_thin, align_center, Font(bold=True)

    for i, dia in enumerate(dias_reporte, 4):
        c = ws.cell(row=2, column=i, value=dia[8:10]+"/"+dia[5:7])
        c.fill, c.border, c.alignment, c.font = fill_yellow, border_thin, align_center, Font(bold=True)

    row_idx = 3
    for doc in docentes:
        ws.cell(row=row_idx, column=1, value=doc.biometric_id).border = border_thin
        ws.cell(row=row_idx, column=2, value=doc.nombre).border = border_thin
        ws.cell(row=row_idx, column=3, value="Docencia").border = border_thin

        for col_idx, dia in enumerate(dias_reporte, 4):
            day_logs = [l for l in all_logs if l.usuario_id == doc.biometric_id and l.fecha.strftime('%Y-%m-%d') == dia]
            
            m_logs = sorted([l for l in day_logs if hora_inicio_m <= l.fecha.time() <= hora_fin_m], key=lambda x: x.fecha)
            t_logs = sorted([l for l in day_logs if hora_inicio_t <= l.fecha.time() <= hora_fin_t], key=lambda x: x.fecha)

            def fmt(logs):
                pre = ""
                if logs[0].origen == "Huella":
                    pre = "(H) "
                elif logs[0].origen == "Asistencia remota":
                    pre = "(W) "

                if len(logs) > 1:
                    return f"{pre}{logs[0].fecha.strftime('%H:%M')}-{logs[-1].fecha.strftime('%H:%M')}"
                return f"{pre}{logs[0].fecha.strftime('%H:%M')}"

            cell_parts = []
            if m_logs:
                cell_parts.append(f"Mañana: {fmt(m_logs)}")
            if t_logs:
                cell_parts.append(f"Tarde: {fmt(t_logs)}")
            
            final_text = "\n".join(cell_parts) or "Sin registros"

            cell = ws.cell(row=row_idx, column=col_idx, value=final_text)
            cell.border, cell.alignment = border_thin, align_center

        row_idx += 1

    ws.column_dimensions['B'].width = 30
    for i in range(4, 4 + len(dias_reporte)): ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="Reporte_ISTAE.xlsx", as_attachment=True)

# --- AJAX Y SEGURIDAD ---
@app.route('/api/logs_admin')
@login_required
def get_logs_admin_json():
    if current_user.rol != 'admin': return jsonify({"error": "No autorizado"}), 403
    
    logs_data = db.session.query(Log, User).outerjoin(User, Log.usuario_id == User.biometric_id).order_by(Log.id.desc()).limit(20).all()
    
    res = []
    for log, user in logs_data:
        log_dict = {
            "fecha": log.fecha.strftime('%Y-%m-%d %H:%M:%S'),
            "nombre": user.nombre if user else f"ID: {log.usuario_id}",
            "tipo_evento": log.tipo_evento,
            "origen": log.origen,
            "lat": log.latitud,
            "lon": log.longitud,
            "foto": log.foto_path,
            "desc": log.descripcion
        }
        res.append(log_dict)
    
    queue_len = Comando.query.filter_by(estado='PENDIENTE').count()
    
    return jsonify(logs=res, queue_length=queue_len)

@app.route('/toggle_permiso/<int:id>', methods=['POST'])
@login_required
def toggle_permiso(id):
    u = db.session.get(User, id)
    if u:
        u.acceso_puerta = 1 if request.json.get('estado') else 0
        db.session.commit()
        return jsonify({"success": True})
    return jsonify({"success": False}), 404

@app.route('/actualizar_password', methods=['POST'])
@login_required
def actualizar_password():
    curr, new = request.form['current_password'], request.form['new_password']
    if check_password_hash(current_user.password, curr):
        current_user.password = generate_password_hash(new, method='pbkdf2:sha256')
        db.session.commit()
        flash('Contraseña actualizada.', 'success')
        return redirect(url_for('index'))
    flash('Contraseña actual incorrecta.', 'danger')
    return redirect(url_for('perfil'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = User.query.filter_by(username=request.form['username']).first()
        if u and check_password_hash(u.password, request.form['password']):
            login_user(u)
            return redirect(url_for('index'))
        flash('Credenciales incorrectas.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
